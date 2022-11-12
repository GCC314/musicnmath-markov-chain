import openpyxl

def Read_from_xlsx(fname, sid = 0):
    wBook = openpyxl.load_workbook(filename=fname)
    sheet = wBook.worksheets[sid]
    lid = 1
    lis = []
    while True:
        sPitch = sheet.cell(row = lid, column = 1).value
        sBeat = sheet.cell(row = lid, column = 2).value
        if(sBeat == None or sBeat == "0"): break
        lis.append((sPitch, int(sBeat)))
        lid += 1
    return lis

def Readall_from_xlsx(fname):
    wBook = openpyxl.load_workbook(filename=fname)
    alis = []
    for sheet in wBook.worksheets:
        lid = 1
        lis = []
        while True:
            sPitch = sheet.cell(row = lid, column = 1).value
            sBeat = sheet.cell(row = lid, column = 2).value
            if(sBeat == None or sBeat == "0"): break
            lis.append((sPitch, int(sBeat)))
            lid += 1
        alis.append(lis)
    return alis

def Save_to_xlsx(lis, fname, sid = 0, isexist = False):
    wBook = openpyxl.Workbook()
    if(isexist): wBook = openpyxl.load_workbook(filename=fname)
    while(len(wBook.worksheets) <= sid): wBook.create_sheet()
    sheet = wBook.worksheets[sid]
    for lid in range(0, len(lis)):
        sheet.cell(lid + 1, 1).value = lis[lid][0]
        sheet.cell(lid + 1, 2).value = lis[lid][1]
    wBook.save(fname)